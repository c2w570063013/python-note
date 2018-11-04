from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
import datetime
import re
import sys
import os
import json
from termcolor import colored

headers = {
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Cache-Control": "max-age=0",
    "Connection": "keep-alive",
    "Content-Type": "application/x-www-form-urlencoded",
    "Host": "exhibitors.electronica.de",
    "Origin": "https://exhibitors.electronica.de",
    "Referer": "https://exhibitors.electronica.de/onlinecatalog/2018/Search_result/",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/69.0.3497.100 Safari/537.36"
}


def parse_excel(file_path):
    try:
        wb = load_workbook(filename=file_path)
        sheet = wb.active
        max_column = sheet.max_column
        max_row = sheet.max_row
        start_row = 7
        start_col = 2
        # first step, add email and domain to the sheet tittle
        sheet.cell(row=start_row - 1, column=max_column + 1).value = 'Email'
        sheet.cell(row=start_row - 1, column=max_column + 2).value = 'Domain'

        # start to iterate the excel and search for the info
        for i in range(start_row, max_row + 1):
            cell_value = sheet.cell(row=i, column=start_col).value.strip()
            # some special string will affect the search result, so we remove them
            process_val = re.sub('Co.{1,4}Ltd.', '', cell_value, flags=re.IGNORECASE)
            process_res = scrape(process_val.strip(), file_path)
            if process_res is None:
                print_r('company:' + process_val.strip() + ' empty', 'warning')
                record_to_log('keyword:' + process_val + '; result:empty', 'empty_' + file_path + '.log')
                continue
            print_r('company:' + process_val + ' found')
            sheet.cell(row=i, column=max_column + 1).value = process_res['email']
            sheet.cell(row=i, column=max_column + 2).value = process_res['domain']
        # save to excel
        wb.save(file_path)

    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        f_name = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        print_r('whoops, error happen! error:' + str(e), 'error', file_path=file_path)
        record_to_log('whoops, error happen! process file name:' + file_path + '; error:' + str(e) + ' ;type:' + str(
            exc_type) + '; file:' + str(f_name) + '; line:' + str(exc_tb.tb_lineno), 'error_' + file_path + '.log')
        pass


def scrape(keyword, file_path):
    try:
        url = 'https://exhibitors.electronica.de/onlinecatalog/2018/Search_result/'
        params = {
            "LNG": 2,
            "nv": 1000,
            "sb_n": "suche",
            "sb_m": 1100,
            "sb_c": 15,
            "pag_styp": 1,
            "sb1_t": "basic",
            "sb1_v": "text",
            "sb1_n": "suche",
            "sb1_searchRequests": "ex,pd",
            "sb1_defaultSearchRequests": "ex,pd",
            "sb2": "ex,pd",
            "sb2_n": "bereiche",
            "sb2_t": "ignoreCondition",
            "sb1": keyword
        }
        res = requests.post(url, data=params, headers=headers, timeout=15)
        if res.status_code != 200:
            raise Exception('network error!', 'status_code' + str(res.status_code))
        soup = BeautifulSoup(res.text, 'html.parser')
        # find the target element
        target = soup.find_all('a', {"class": 'exd_navfont'})
        if len(target) == 0:
            # check if exists multi results
            target2 = soup.find_all('a', {'class': 'ex_font'})
            # print(target2[0]['href'])
            if len(target2) == 0:
                return None
            # we just use the first links to search for the info, not consider the rest of results
            url2 = target2[0]['href'].strip()
            return scrape_2(url2, keyword, file_path)
        return extract_data(soup, keyword, file_path)
    except Exception as e:
        print_r(str(e), 'error', keyword=keyword, file_path=file_path)
        record_to_log('keyword:' + keyword + '; error:' + str(e), 'error_' + file_path + '.log')


def scrape_2(url, keyword, file_path):
    try:
        res = requests.get(url, headers=headers, timeout=15)
        if res.status_code != 200:
            raise Exception('network error!', 'status_code' + str(res.status_code))
        soup = BeautifulSoup(res.text, 'html.parser')
        return extract_data(soup, keyword, file_path)
    except Exception as e:
        record_to_log('keyword:' + keyword + '; error:' + str(e), 'error_' + file_path + '.log')


def print_r(words, level='success', keyword=None, file_path=None):
    color = 'green'
    if level == 'warning':
        color = 'yellow'
    if level == 'error':
        color = 'red'
    external_content = ''
    if keyword is not None:
        external_content = external_content + ' keyword:' + keyword + ';'
    if file_path is not None:
        external_content = external_content + ' file path:' + file_path
    print(colored(level + ': ' + words + ';' + external_content, color))


def extract_data(soup, keyword, file_path):
    try:
        # find the target element
        target = soup.find_all('a', {"class": 'exd_navfont'})

        res_list = {}
        for a in target:
            a_val = a.text.strip()
            # determine if value is email
            if '@' in a_val:
                res_list['email'] = a_val
                continue
            res_list['domain'] = a_val + ' '
        # if no specific key in return value, set default '' to it
        if 'email' not in res_list:
            res_list['email'] = ''
        if 'domain' not in res_list:
            res_list['domain'] = ''
        # remove space from domain
        res_list['domain'] = res_list['domain'].strip()
        record_to_log('keyword:' + keyword + '; result:' + json.dumps(res_list), 'success_' + file_path + '.log')
        return res_list
    except Exception as e:
        print_r(str(e), 'error', keyword=keyword, file_path=file_path)
        record_to_log('keyword:' + keyword + '; error:' + str(e), 'error_' + file_path + '.log')


def record_to_log(content, file='log.txt'):
    time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    f = open(file, 'a')
    f.write('[' + time + '] ' + content + "\n")


if __name__ == "__main__":
    files = ['excel/(electronica experience).xlsx', 'excel/(Embedded).xlsx', 'excel/(New Exhibitor).xlsx',
             'excel/SEMICON Europa.xlsx']
    # get excel file from terminal args
    if len(sys.argv) > 1:
        files = sys.argv[1:]
    for file_ in files:
        print_r('-----------------start process ' + file_ + ' -----------------')
        parse_excel(file_)
        print_r('-----------------process ' + file_ + ' done!-----------------')
